import sys
import json
import re
import os

# Status ranking dictionary
# Rank 1 = NOT_STARTED (Lowest Rank)
# Rank 13 = AC_CREATED (Highest Winning Rank)
# Higher rank number = Better status. Winner = bank with HIGHEST rank number.
STATUS_RANKING = {
    'NOT_STARTED': 1,
    'NOT_APPLICABLE': 2,
    'REJECTED': 3,
    'IN_PROGRESS': 4,
    'DOC_UPLOAD_PENDING': 5,
    'DOC_UPLOADED': 6,
    'KYC_PENDING': 7,
    'DOC_REUPLOAD_REQUIRED': 8,
    'KYC_DONE': 9,
    'SUBMITTED': 10,
    'SUBMITTED_OTP_VERIFIED': 11,
    'VKYC_PENDING': 12,
    'AC_CREATED': 13
}

def clean_key_val(val):
    if val is None:
        return ''
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return re.sub(r'[^a-z0-9]', '', s.lower())

def get_status_rank(status_str):
    if not status_str:
        return 0
    s = str(status_str).strip().upper()
    clean = re.sub(r'[^A-Z0-9]', '', s)

    if 'ACCREATED' in clean or 'ACCOUNTCREATED' in clean:
        return 13
    if 'VKYCPENDING' in clean or 'VKYC' in clean:
        return 12
    if 'OTPVERIFIED' in clean or 'SUBMITTEDOTP' in clean:
        return 11
    if 'SUBMITTED' in clean:
        return 10
    if 'KYCDONE' in clean or 'KYCCOMPLETED' in clean:
        return 9
    if 'DOCREUPLOAD' in clean or 'REUPLOAD' in clean:
        return 8
    if 'KYCPENDING' in clean:
        return 7
    if 'DOCUPLOADED' in clean or 'DOCUMENTUPLOADED' in clean:
        return 6
    if 'DOCUPLOADPENDING' in clean or 'DOCPENDING' in clean or 'DOCUMENTPENDING' in clean:
        return 5
    if 'INPROGRESS' in clean or 'PROCESSING' in clean:
        return 4
    if 'REJECT' in clean or 'DECLINE' in clean or 'CANCEL' in clean:
        return 3
    if 'NOTAPPLICABLE' in clean or 'NA' in clean:
        return 2
    if 'NOTSTARTED' in clean:
        return 1

    clean_u = re.sub(r'[\s-]+', '_', s)
    return STATUS_RANKING.get(clean_u, 0)

def extract_urn(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    m = re.search(r'FM\d{4}[A-Z0-9]\d{6,12}', s, re.IGNORECASE) or \
        re.search(r'FM\d{4}\d{6,12}', s, re.IGNORECASE) or \
        re.search(r'FM[0-9A-Z]{8,18}', s, re.IGNORECASE)
    return m.group(0).upper() if m else None

def parse_xlsx_fast(file_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        
        def read_sheet_rows(sheet_name):
            if sheet_name not in sheet_names:
                return []
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers_raw = next(rows_iter)
            except StopIteration:
                return []
            if headers_raw and any(headers_raw):
                headers = [str(h).strip() if h is not None else f'col_{idx}' for idx, h in enumerate(headers_raw)]
            else:
                return []
            
            result = []
            for r in rows_iter:
                if not r or not any(r):
                    continue
                row_dict = {}
                for idx, val in enumerate(r):
                    if idx < len(headers):
                        row_dict[headers[idx]] = '' if val is None else str(val).strip()
                result.append(row_dict)
            return result

        yes_name = next((s for s in sheet_names if 'yes' in s.lower()), None)
        au_name = next((s for s in sheet_names if 'au' in s.lower()), None)
        pnb_name = next((s for s in sheet_names if 'pnb' in s.lower()), None)

        if not yes_name and not au_name and not pnb_name:
            return {"error": "No valid KIWI sheets (YES, AU, PNB) found in uploaded Excel file."}

        yes_rows = read_sheet_rows(yes_name) if yes_name else []
        au_rows = read_sheet_rows(au_name) if au_name else []
        pnb_rows = read_sheet_rows(pnb_name) if pnb_name else []

        wb.close()
        return process_kiwi_rows(yes_rows, au_rows, pnb_rows)

    except Exception as e:
        return {"error": f"Failed to parse Excel file: {str(e)}"}

def get_row_value(row_dict, target_name):
    if not row_dict:
        return ''
    clean_target = re.sub(r'[^a-z0-9]', '', target_name.lower())
    for k, v in row_dict.items():
        clean_k = re.sub(r'[^a-z0-9]', '', k.lower())
        if clean_k == clean_target:
            return str(v).strip()
    return ''

def get_state_val(row_dict):
    if not row_dict:
        return ''
    for target in ('current_state', 'current_status', 'app_state', 'app_status', 'status'):
        val = get_row_value(row_dict, target)
        if val and val.lower() not in ('state', 'customer_state'):
            return val
    # Fallback scan keys
    for k, v in row_dict.items():
        lk = k.lower().strip()
        if lk not in ('state', 'customer_state') and ('state' in lk or 'status' in lk):
            if v:
                return str(v).strip()
    return ''

def get_user_id_val(row_dict):
    if not row_dict:
        return ''
    # Prioritize exact lower 'user_id' over upper 'USER_ID' if present
    for k, v in row_dict.items():
        if k.strip() == 'user_id' and v:
            return str(v).strip()
    for k, v in row_dict.items():
        if k.strip().lower() == 'user_id' and v:
            return str(v).strip()
    for target in ('user_id', 'userid', 'user_identifier', 'useridentifier', 'user'):
        val = get_row_value(row_dict, target)
        if val:
            return val
    return ''

def extract_all_row_keys(row_dict):
    """
    Returns a set of cleaned key values (user_id, URN, app_id) for robust cross-sheet lookup.
    """
    keys = set()
    if not row_dict:
        return keys

    # 1. User ID
    uid = get_user_id_val(row_dict)
    if uid:
        c_uid = clean_key_val(uid)
        if c_uid:
            keys.add(c_uid)

    # 2. Application ID (bank 2)
    app_id = get_row_value(row_dict, 'application_id_bank_2') or get_row_value(row_dict, 'application_id')
    if app_id:
        c_app = clean_key_val(app_id)
        if c_app:
            keys.add(c_app)

    # 3. URN from content/Reference/URN
    content_val = get_row_value(row_dict, 'content') or get_row_value(row_dict, 'contant') or get_row_value(row_dict, 'reference') or get_row_value(row_dict, 'urn')
    urn = extract_urn(content_val)
    if not urn:
        # scan values for URN
        for v in row_dict.values():
            urn = extract_urn(v)
            if urn:
                break
    if urn:
        keys.add(urn.upper())

    return keys

def index_sheet_rows(rows_list):
    """
    Indexes rows by all possible keys (user_id, application_id_bank_2, URN)
    Returns: (map_by_key, list_of_rows_with_urn)
    """
    row_map = {}
    rows_with_urn = []

    for r in rows_list:
        # Pre-compute state
        r['_state'] = get_state_val(r)
        
        # Check URN
        content_val = get_row_value(r, 'content') or get_row_value(r, 'contant') or get_row_value(r, 'reference') or get_row_value(r, 'urn')
        urn = extract_urn(content_val)
        if not urn:
            for v in r.values():
                urn = extract_urn(v)
                if urn:
                    break
        if urn:
            r['_urn'] = urn.upper()
            rows_with_urn.append(r)

        # Index under all identifiers
        row_keys = extract_all_row_keys(r)
        for k in row_keys:
            if k not in row_map:
                row_map[k] = r

    return row_map, rows_with_urn

def process_kiwi_rows(yes_rows, au_rows, pnb_rows):
    yes_map, yes_urn_rows = index_sheet_rows(yes_rows)
    au_map, au_urn_rows = index_sheet_rows(au_rows)
    pnb_map, pnb_urn_rows = index_sheet_rows(pnb_rows)

    # Collect all unique URNs across all 3 sheets (YES primary, then AU, PNB)
    urn_to_lead = {}

    def add_lead_sources(rows_list, default_bank):
        for r in rows_list:
            urn = r.get('_urn')
            if urn and urn not in urn_to_lead:
                urn_to_lead[urn] = {
                    'urn': urn,
                    'primary_row': r,
                    'primary_bank': default_bank
                }

    add_lead_sources(yes_urn_rows, 'YES')
    add_lead_sources(au_urn_rows, 'AU')
    add_lead_sources(pnb_urn_rows, 'PNB')

    parsed_rows = []

    for urn, lead_info in urn_to_lead.items():
        p_row = lead_info['primary_row']
        p_bank = lead_info['primary_bank']

        # Extract all keys for this lead from primary_row
        lead_keys = extract_all_row_keys(p_row)

        def find_candidate_row(sheet_map):
            for k in lead_keys:
                if k in sheet_map:
                    return sheet_map[k]
            return None

        cand_yes = find_candidate_row(yes_map) or (p_row if p_bank == 'YES' else None)
        cand_au = find_candidate_row(au_map) or (p_row if p_bank == 'AU' else None)
        cand_pnb = find_candidate_row(pnb_map) or (p_row if p_bank == 'PNB' else None)

        yes_state = str(cand_yes.get('_state') if cand_yes else '').strip()
        au_state = str(cand_au.get('_state') if cand_au else '').strip()
        pnb_state = str(cand_pnb.get('_state') if cand_pnb else '').strip()

        yes_rank = get_status_rank(yes_state)
        au_rank = get_status_rank(au_state)
        pnb_rank = get_status_rank(pnb_state)

        # Compare ranks: Higher rank number wins (13 = AC_CREATED best, 1 = NOT_STARTED worst)
        winning_bank = p_bank
        winning_row = p_row
        winning_state = str(p_row.get('_state') or '').strip()
        best_rank = get_status_rank(winning_state)

        if yes_rank > best_rank and cand_yes:
            best_rank = yes_rank
            winning_bank = 'YES'
            winning_row = cand_yes
            winning_state = yes_state

        if au_rank > best_rank and cand_au:
            best_rank = au_rank
            winning_bank = 'AU'
            winning_row = cand_au
            winning_state = au_state

        if pnb_rank > best_rank and cand_pnb:
            best_rank = pnb_rank
            winning_bank = 'PNB'
            winning_row = cand_pnb
            winning_state = pnb_state

        wr = winning_row or p_row
        uid = get_user_id_val(wr) or get_user_id_val(p_row)

        parsed_rows.append({
            'content': urn,
            'registration': get_row_value(wr, 'registration'),
            'pan_submit': get_row_value(wr, 'Pan_Submit') or get_row_value(wr, 'pan_submit'),
            'form_fetch': get_row_value(wr, 'Form_Fetch') or get_row_value(wr, 'form_fetch'),
            'form_submit': get_row_value(wr, 'Form_Submit') or get_row_value(wr, 'form_submit'),
            'ipa': get_row_value(wr, 'IPA') or get_row_value(wr, 'ipa'),
            'card_created': get_row_value(wr, 'Card_Created') or get_row_value(wr, 'card_created'),
            'vkyc': get_row_value(wr, 'VKYC') or get_row_value(wr, 'vkyc'),
            'current_state': winning_state,
            'reject_reason': get_row_value(wr, 'reject_reason'),
            'application_id_bank_2': get_row_value(wr, 'application_id_bank_2'),
            'first_txn': get_row_value(wr, 'First_txn') or get_row_value(wr, 'first_txn'),
            'APPLICATION_REFERENCE_NUMBER': urn,
            'current_status': winning_state,
            'final_decision': winning_state,
            'kiwi_winning_bank': winning_bank,
            'kiwi_user_id': uid,
            'kiwi_yes_status': yes_state,
            'kiwi_au_status': au_state,
            'kiwi_pnb_status': pnb_state,
            '_extractedUrn': urn,
            'yes_rank': yes_rank,
            'au_rank': au_rank,
            'pnb_rank': pnb_rank,
            'status_rank': best_rank
        })

    return {
        "success": True,
        "parsedRows": parsed_rows,
        "totalRows": len(parsed_rows),
        "yesRowsCount": len(yes_rows),
        "auRowsCount": len(au_rows),
        "pnbRowsCount": len(pnb_rows)
    }

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No file path provided."}))
        sys.exit(1)
    file_path = sys.argv[1]
    res = parse_xlsx_fast(file_path)
    print(json.dumps(res))
